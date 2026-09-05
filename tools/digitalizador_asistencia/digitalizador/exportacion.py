"""Salidas: CSV, JSON, XLSX e informe de revisión."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Dict, List, Sequence

from .modelos import Hoja

COLUMNAS_BASE = [
    "archivo", "pagina", "fila", "numero", "nombre", "documento",
]
COLUMNAS_FINALES = [
    "confianza", "revisar", "motivos", "nombre_leido", "padron_estado", "padron_puntaje",
]


def _filas(hojas: Sequence[Hoja]) -> List[Dict[str, Any]]:
    filas: List[Dict[str, Any]] = []
    for hoja in hojas:
        for registro in hoja.registros:
            fila = {"archivo": hoja.origen, "pagina": hoja.pagina}
            fila.update(registro.a_dict())
            filas.append(fila)
    return filas


def _cabecera(hojas: Sequence[Hoja], filas: Sequence[Dict[str, Any]]) -> List[str]:
    columnas: List[str] = list(COLUMNAS_BASE)
    for hoja in hojas:
        for columna in hoja.columnas:
            if columna not in columnas:
                columnas.append(columna)
    for fila in filas:
        for clave in fila:
            if clave not in columnas and clave not in COLUMNAS_FINALES:
                columnas.append(clave)
    return columnas + COLUMNAS_FINALES


def exportar_csv(hojas: Sequence[Hoja], destino: Path | str) -> Path:
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    filas = _filas(hojas)
    cabecera = _cabecera(hojas, filas)
    with destino.open("w", encoding="utf-8-sig", newline="") as manejador:
        escritor = csv.DictWriter(manejador, fieldnames=cabecera, extrasaction="ignore")
        escritor.writeheader()
        for fila in filas:
            escritor.writerow({clave: fila.get(clave, "") for clave in cabecera})
    return destino


def exportar_json(hojas: Sequence[Hoja], destino: Path | str) -> Path:
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    datos = {
        "resumen": [hoja.resumen() for hoja in hojas],
        "hojas": [
            {
                "archivo": hoja.origen,
                "pagina": hoja.pagina,
                "columnas": hoja.columnas,
                "encabezados": hoja.encabezados,
                "metadatos": hoja.metadatos,
                "registros": [registro.a_dict() for registro in hoja.registros],
            }
            for hoja in hojas
        ],
    }
    destino.write_text(
        json.dumps(datos, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    return destino


def exportar_xlsx(hojas: Sequence[Hoja], destino: Path | str) -> Path:
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Font, PatternFill  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Falta 'openpyxl'. Instala: pip install openpyxl") from exc

    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    filas = _filas(hojas)
    cabecera = _cabecera(hojas, filas)

    libro = Workbook()
    hoja_calculo = libro.active
    hoja_calculo.title = "Asistencia"
    hoja_calculo.append(cabecera)
    for celda in hoja_calculo[1]:
        celda.font = Font(bold=True)

    resalte = PatternFill("solid", fgColor="FFF3CD")
    for fila in filas:
        hoja_calculo.append([fila.get(clave, "") for clave in cabecera])
        if fila.get("revisar"):
            for celda in hoja_calculo[hoja_calculo.max_row]:
                celda.fill = resalte

    for indice, nombre in enumerate(cabecera, start=1):
        ancho = max(len(str(nombre)), *(len(str(f.get(nombre, ""))) for f in filas)) if filas else len(str(nombre))
        hoja_calculo.column_dimensions[hoja_calculo.cell(row=1, column=indice).column_letter].width = min(
            max(ancho + 2, 8), 48
        )
    hoja_calculo.freeze_panes = "A2"
    libro.save(destino)
    return destino


def informe(hojas: Sequence[Hoja]) -> str:
    """Informe de texto con el resumen y las filas que hay que revisar."""
    lineas: List[str] = ["=== Digitalización de listados de asistencia ==="]
    total = revisar = relecturas = con_error = 0
    for hoja in hojas:
        resumen = hoja.resumen()
        total += resumen["filas"]
        revisar += resumen["a_revisar"]
        relecturas += sum(1 for registro in hoja.registros if registro.relectura)
        lineas.append(
            f"\n{resumen['origen']} (página {resumen['pagina']}): "
            f"{resumen['filas']} filas, {resumen['presentes']} presentes, "
            f"{resumen['ausentes']} ausentes, {resumen['marcas_dudosas']} marcas dudosas"
        )
        if hoja.metadatos.get("error"):
            con_error += 1
            lineas.append(f"  ERROR: {hoja.metadatos['error']}")
        if hoja.metadatos.get("aviso"):
            lineas.append(f"  aviso: {hoja.metadatos['aviso']}")
        if hoja.metadatos.get("rotacion_corregida"):
            lineas.append(f"  se giró la página {hoja.metadatos['rotacion_corregida']}°")
        segunda = hoja.metadatos.get("segunda_opinion")
        if segunda:
            lineas.append(
                f"  segunda lectura: {segunda.get('consultadas', 0)} filas releídas, "
                f"{segunda.get('actualizadas', 0)} corregidas, "
                f"{segunda.get('ilegibles', 0)} siguen ilegibles"
            )
        for registro in hoja.a_revisar:
            nombre = registro.nombre.texto or "(sin nombre)"
            lineas.append(f"  · fila {registro.fila}: {nombre} — {'; '.join(registro.motivos)}")

    lineas.append(f"\nTotal: {total} filas, {revisar} requieren revisión manual.")
    if relecturas:
        lineas.append(f"Filas mejoradas en la segunda lectura: {relecturas}.")
    if con_error:
        lineas.append(f"Páginas con error: {con_error} (ver el detalle arriba).")
    if total:
        lineas.append(f"Fiabilidad automática: {(total - revisar) / total:.0%} de las filas.")
    return "\n".join(lineas)


def exportar(
    hojas: Sequence[Hoja], destino: Path | str, formato: str = "csv"
) -> List[Path]:
    """Escribe la salida en el/los formatos pedidos ('csv', 'json', 'xlsx' o 'todos')."""
    destino = Path(destino)
    formatos = ["csv", "json", "xlsx"] if formato == "todos" else [formato]
    escritos: List[Path] = []
    for nombre in formatos:
        ruta = destino if destino.suffix.lower() == f".{nombre}" else destino.with_suffix(f".{nombre}")
        if nombre == "csv":
            escritos.append(exportar_csv(hojas, ruta))
        elif nombre == "json":
            escritos.append(exportar_json(hojas, ruta))
        elif nombre == "xlsx":
            escritos.append(exportar_xlsx(hojas, ruta))
        else:
            raise ValueError(f"Formato desconocido: {nombre}")
    return escritos
